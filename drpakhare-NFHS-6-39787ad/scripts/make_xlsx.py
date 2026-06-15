import pandas as pd, numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
m=pd.read_csv("/sessions/sleepy-bold-wozniak/mnt/outputs/master_long3.csv")

ROUND_ORDER=["NFHS-3","NFHS-4","NFHS-5","NFHS-6"]
m["round"]=pd.Categorical(m["round"],ROUND_ORDER,ordered=True)

# ---------- Master_Long ----------
master=m[["state_std","round","year","area","section","indicator","canonical",
          "value","value_raw","data_flag","ind_no","source"]].copy()
master=master.rename(columns={"state_std":"state","canonical":"harmonized_indicator"})
master=master.sort_values(["state","round","section","indicator","area"]).reset_index(drop=True)

# ---------- Indicator_Dictionary ----------
dic=(m.groupby(["round","section","indicator","canonical"],dropna=False,observed=True)
       .agg(geographies_with_data=("value",lambda s:s.notna().sum()),
            areas=("area","nunique")).reset_index()
       .rename(columns={"canonical":"harmonized_indicator"}))
dic=dic.sort_values(["round","section","indicator"])

# ---------- Harmonized_Coverage ----------
cw_names=[]
import re
for line in open("/sessions/sleepy-bold-wozniak/mnt/outputs/crosswalk.py"):
    mm=re.match(r'\s*\("([^"]+)",',line)
    if mm and mm.group(1) not in cw_names: cw_names.append(mm.group(1))
cov_rows=[]
for name in cw_names:
    sub=m[m["canonical"]==name]
    d={"harmonized_indicator":name}
    sec=sub["section"].dropna()
    d["example_section"]=sec.iloc[0] if len(sec) else ""
    for r in ROUND_ORDER:
        d[r]="Y" if sub[sub["round"]==r]["value"].notna().any() else ""
    d["rounds_covered"]=sum(1 for r in ROUND_ORDER if d[r]=="Y")
    cov_rows.append(d)
cov=pd.DataFrame(cov_rows).sort_values(["rounds_covered","harmonized_indicator"],ascending=[False,True])

# ---------- Trend_Total (harmonized, Total area) ----------
t=m[(m["area"]=="Total")&(m["canonical"].notna())].copy()
trend=t.pivot_table(index=["state_std","canonical"],columns="round",
                    values="value",aggfunc="mean",observed=True).reset_index()
trend=trend.rename(columns={"state_std":"state","canonical":"harmonized_indicator"})
trend=trend.sort_values(["harmonized_indicator","state"])

# ---------- README ----------
readme=[
("NFHS Combined Database — India & State/UT Fact Sheets",""),
("Built",pd.Timestamp.today().strftime("%Y-%m-%d")),
("",""),
("ROUNDS & SOURCES",""),
("NFHS-6 (2023-24)","Extracted from official Fact Sheets PDF (May 2026). Urban/Rural/Total."),
("NFHS-5 (2019-21)","From NFHS_5_Factsheets_Data.xls. Urban/Rural/Total."),
("NFHS-4 (2015-16)","From 'NFHS-4 and 3 Fact Sheets Sparse.xlsx'. Urban/Rural/Total."),
("NFHS-3 (2005-06)","From same sparse file. Mostly Total only; many indicators 'NA'."),
("",""),
("SHEETS",""),
("Master_Long","Tidy data: one row per state x round x area x indicator. The full database."),
("Indicator_Dictionary","Every distinct indicator per round, with its harmonized name where mapped."),
("Harmonized_Coverage","84 curated comparable indicators x which rounds report them (Y)."),
("Trend_Total","Pivot of harmonized indicators (Total area) across rounds — for trend analysis."),
("",""),
("KEY NOTES",""),
("Geography","36 fact sheets in NFHS-6 (India + 28 states + UTs). MANIPUR is ABSENT from NFHS-6."),
("","NFHS-4/3 list Dadra & Nagar Haveli and Daman & Diu SEPARATELY (pre-2020 merger);"),
("","NFHS-5/6 report them as one merged UT — these do not align across rounds."),
("","Ladakh exists only in NFHS-5/6 (carved from J&K in 2019)."),
("State names","Standardised (e.g. Maharastra->Maharashtra, Chattisgarh->Chhattisgarh, Delhi->NCT of Delhi)."),
("value","Numeric value. 'value_raw' keeps the original cell text (incl. NA / footnote markers)."),
("Harmonisation","'harmonized_indicator' aligns equivalently-defined indicators across rounds despite"),
("","wording changes. 52 indicators span all 4 rounds, 23 span 3, 9 span 2."),
("Caution","Indicator DEFINITIONS sometimes changed across rounds (age groups, denominators,"),
("","vaccine schedules). Treat cross-round comparisons as indicative; verify definitions in"),
("","the source fact sheet footnotes before publication."),
]
readme_df=pd.DataFrame(readme,columns=["Item","Detail"])

OUT="/sessions/sleepy-bold-wozniak/mnt/NFHS-6/NFHS_Combined_Database_3to6.xlsx"
with pd.ExcelWriter(OUT,engine="openpyxl") as xw:
    readme_df.to_excel(xw,"README",index=False)
    master.to_excel(xw,"Master_Long",index=False)
    dic.to_excel(xw,"Indicator_Dictionary",index=False)
    cov.to_excel(xw,"Harmonized_Coverage",index=False)
    trend.to_excel(xw,"Trend_Total",index=False)

    wb=xw.book
    hdr_fill=PatternFill("solid",fgColor="1F4E78"); hdr_font=Font(color="FFFFFF",bold=True)
    for sh in wb.worksheets:
        sh.freeze_panes="A2"
        if sh.title!="README":
            sh.auto_filter.ref=sh.dimensions
        for c in sh[1]:
            c.fill=hdr_fill; c.font=hdr_font; c.alignment=Alignment(vertical="center",wrap_text=True)
        sh.row_dimensions[1].height=28
        # column widths
        for col in sh.columns:
            letter=get_column_letter(col[0].column)
            maxlen=max((len(str(c.value)) for c in col[:200] if c.value is not None),default=10)
            sh.column_dimensions[letter].width=min(max(maxlen+2,10),60)
    # README wider detail
    wb["README"].column_dimensions["A"].width=22
    wb["README"].column_dimensions["B"].width=95
    wb["README"]["A1"].font=Font(bold=True,size=14,color="1F4E78")
print("WROTE",OUT)
print("Master_Long rows:",len(master))
print("Dictionary rows:",len(dic))
print("Coverage indicators:",len(cov))
print("Trend rows:",len(trend),"cols:",list(trend.columns))
